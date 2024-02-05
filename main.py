import openpyxl
from pathlib import Path
import csv
import sys
import os
import shutil

# 団体のdict
groups = {}

# 代表者
representative = []
# 会計長
accountant = []
# 総務
general_affairs = []
# 副代表者
deputy_representative = []
# 副会計長
deputy_accountant = []


def load_group_data(file_name):
    wb = openpyxl.load_workbook(filename=file_name, data_only=True)
    ws = wb["Sheet1"]
    for row in ws.rows:
        groups[row[0].value] = row[1].value+row[3].value

def read_role(file_name):
    wb = openpyxl.load_workbook(filename=file_name, data_only=True)
    ws = wb["Sheet1"]
    for row in ws.rows:
        if (row[4].value == "部長・主将" or
                row[4].value == "会計長" or
                row[4].value == "総務" or
                row[4].value == "副部長・副主将" or
                row[4].value == "副会計長"
        ):
            group_name = groups.get(ws["A1"].value)
            if group_name is None:
                group_name = ws["D1"].value

            tel_num = row[5].value

            print(file_name)
            if tel_num and len(str(tel_num)) < 13:
                number_list = list(str(row[5].value))
                number_list.insert(3, "-")
                number_list.insert(8, "-")
                tel_num = ''.join(number_list)

            row_data = [
                row[2].value,
                "",
                str(row[3].value).upper(),
                tel_num,
                group_name
            ]

            print(row[4].value)
            if row[4].value == "部長・主将":
                representative.append(row_data)
            elif row[4].value == "会計長":
                accountant.append(row_data)
            elif row[4].value == "総務":
                general_affairs.append(row_data)
            elif row[4].value == "副部長・副主将":
                deputy_representative.append(row_data)
            elif row[4].value == "副会計長":
                deputy_accountant.append(row_data)


def create_output_file():
    leaders_list = [
        ["代表者", representative],
        ["会計長", accountant],
        ["総務", general_affairs],
        ["副代表者", deputy_representative],
        ["副会計長", deputy_accountant]
    ]

    wb = openpyxl.Workbook()

    for leaders in leaders_list:
        representative_sheet = wb.create_sheet(title=leaders[0])
        i = 0
        for row in representative_sheet.iter_rows(min_row=1, min_col=1,
                                                  max_row=len(leaders[1]),
                                                  max_col=len(leaders[1][0])):
            j = 0
            for cell in row:
                # print(leaders[1][i][j])
                cell.value = leaders[1][i][j]
                j += 1
            i += 1

    wb.save('output_file.xlsx')


load_group_data("./group_list.xlsx")

dir_path = "member_list"

files = [
    os.path.join(dir_path, f) for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))
]

for file_path in files:
    read_role(file_path)

create_output_file()